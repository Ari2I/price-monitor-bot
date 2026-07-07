"""
Экспорт отчёта по отслеживаемым товарам в файл Excel.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Валюта по умолчанию для отображения, если её не удалось определить
# автоматически — используется только для показа в отчёте, а не как
# гарантированный факт о товаре.
DEFAULT_CURRENCY_LABEL = "₽"

COLUMN_TITLES = [
    "Название",
    "Ссылка",
    "Текущая цена",
    "Валюта",
    "Предыдущая цена",
    "Изменение",
    "Изменение, %",
    "Последняя проверка",
]


@dataclass
class ProductReportRow:
    """Одна строка отчёта по товару."""

    name: str
    url: str
    current_price: Optional[float]
    currency: Optional[str]
    previous_price: Optional[float]
    checked_at: Optional[datetime]

    @property
    def change(self) -> Optional[float]:
        if self.current_price is None or self.previous_price is None:
            return None
        return round(self.current_price - self.previous_price, 2)

    @property
    def change_percent(self) -> Optional[float]:
        if not self.change or not self.previous_price:
            return 0.0 if self.change == 0 else None
        return round((self.change / self.previous_price) * 100, 2)


def export_report_to_excel(
    rows: List[ProductReportRow], file_path: str
) -> None:
    """Сохраняет отчёт по товарам в XLSX-файл."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Мониторинг цен"

    sheet.append(COLUMN_TITLES)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(
            [
                row.name,
                row.url,
                row.current_price,
                row.currency or DEFAULT_CURRENCY_LABEL,
                row.previous_price,
                row.change,
                row.change_percent,
                row.checked_at.strftime("%Y-%m-%d %H:%M")
                if row.checked_at
                else None,
            ]
        )

    for index, _ in enumerate(COLUMN_TITLES, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = 22

    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(file_path)
