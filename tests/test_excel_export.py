"""Тесты для export.excel_export."""

import os
import tempfile
import unittest
from datetime import datetime

from openpyxl import load_workbook

from export.excel_export import ProductReportRow, export_report_to_excel


class ExcelExportTest(unittest.TestCase):
    def test_export_creates_file_with_expected_rows(self):
        rows = [
            ProductReportRow(
                name="Товар 1",
                url="https://example.com/1",
                current_price=1000.0,
                currency="USD",
                previous_price=900.0,
                checked_at=datetime(2026, 6, 30, 12, 0),
            ),
            ProductReportRow(
                name="Товар 2",
                url="https://example.com/2",
                current_price=None,
                currency=None,
                previous_price=None,
                checked_at=None,
            ),
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = os.path.join(tmp_dir, "report.xlsx")
            export_report_to_excel(rows, file_path)

            self.assertTrue(os.path.exists(file_path))

            workbook = load_workbook(file_path)
            sheet = workbook.active

            self.assertEqual(sheet.cell(row=1, column=1).value, "Название")
            self.assertEqual(sheet.cell(row=1, column=4).value, "Валюта")
            self.assertEqual(sheet.cell(row=2, column=1).value, "Товар 1")
            self.assertEqual(sheet.cell(row=2, column=3).value, 1000.0)
            self.assertEqual(sheet.cell(row=2, column=4).value, "USD")
            self.assertEqual(sheet.cell(row=2, column=6).value, 100.0)

    def test_missing_currency_falls_back_to_default_label(self):
        rows = [
            ProductReportRow(
                name="Товар без валюты",
                url="https://example.com/3",
                current_price=500.0,
                currency=None,
                previous_price=None,
                checked_at=None,
            ),
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = os.path.join(tmp_dir, "report.xlsx")
            export_report_to_excel(rows, file_path)

            workbook = load_workbook(file_path)
            sheet = workbook.active
            self.assertEqual(sheet.cell(row=2, column=4).value, "₽")

    def test_change_and_change_percent_calculations(self):
        row = ProductReportRow(
            name="Товар",
            url="https://example.com",
            current_price=110.0,
            currency="RUB",
            previous_price=100.0,
            checked_at=datetime.now(),
        )
        self.assertEqual(row.change, 10.0)
        self.assertEqual(row.change_percent, 10.0)

    def test_change_is_none_without_previous_price(self):
        row = ProductReportRow(
            name="Товар",
            url="https://example.com",
            current_price=110.0,
            currency="RUB",
            previous_price=None,
            checked_at=datetime.now(),
        )
        self.assertIsNone(row.change)


if __name__ == "__main__":
    unittest.main()
